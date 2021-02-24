import os
import shutil
import tkinter as tk
from tkinter import filedialog
from tkinter.filedialog import askopenfilename 
from tkinter import messagebox
from tkinter import Button
import openpyxl
from openpyxl import load_workbook
from pathlib import Path

list_PN_to_get = []

#base = 'C:\\Users\\Michael\\Desktop\\work\\Elbit files\\Reciever collector\\TEST_DRIVE'
base = 'R:\\Receipts\\DAP RECEIVERS '
print("i think the base location is: " + base)

def folder_out_func():
    
    global output_loc

    root.directory = filedialog.askdirectory()
    
    output_loc = root.directory
    print("I think the output loc is: " + output_loc)
    
    #out_path = 'C:\\Users\\Michael\\Desktop\\work\\Elbit files\\Reciever collector\\Output'

    tk.Label(root, text=output_loc).grid(row = 2, column = 2, sticky = 'W')

    #but_coll.grid(row = 3, column=1, sticky='W')

    #run_collection(list_PN_to_get, output_loc)

    return output_loc


def get_xl_list():
    xl_path = askopenfilename()
    #print(xl_path)
    
    #out_path = 'C:\\Users\\Michael\\Desktop\\work\\Elbit files\\Reciever collector\\Output'
    
    wb = load_workbook(xl_path)
    ws = wb['Sheet1']

    for i in range(1,len(ws['A'])+1):
        name_item = 'A' + str(i)
        PN_from_xl = ws[name_item].value

        list_PN_to_get.append(PN_from_xl)

    #print(list_PN_to_get)

    if list_PN_to_get != []:
        tk.Label(root, text='This looks ok!').grid(row = 1, column = 2, sticky = 'W')

    #run_collection(list_PN_to_get, output_loc)

    print("these are the PN's I am trying to get:")
    print(*list_PN_to_get, sep = ", ")

    return list_PN_to_get


def run_collection():

    #print(list_PN_to_get)
    #print(output_loc)

    #may need to change years
    start = 2019
    end = 2020

    list_PN_found = []

    for PN in list_PN_to_get:

        for year in range(start, (end+1)):

            path = os.path.join(base, str(year))

            sorted_file_names = [os.path.join(path, f) for f in os.listdir(path)] # add path to each file
            sorted_file_names.sort(key=lambda x: os.path.getmtime(x))

            fixed_list = sorted_file_names[::-1]

            #print(fixed_list)

            for iter in fixed_list:
                

                if PN in iter:

                    print("I think i have found: " + PN + " at: " + iter)
                    sep = '\\'
                    this_file_name = iter.rsplit(sep, 1)[-1]

                    pn_name = str(PN)

                    first = os.path.join(output_loc, pn_name)

                    os.makedirs(first)

                    full_path_out = os.path.join(first, this_file_name)

                    print("I will try to put it here: " + full_path_out)

                    shutil.copy2(iter, full_path_out)

                    list_PN_found.append(PN)

                    break



root = tk.Tk()

#set window geometry
root.geometry("600x200")

#img = Image.open(r'C:\Users\Michael\Desktop\Elbit files\Receipt Reader\Elbit_Logo.png')
#img = img.resize((192,66), Image.ANTIALIAS)
#photoImg = itk.PhotoImage(img)

#background_label = tk.Label(root, image=photoImg)
#background_label.place(relx = 1, rely = 1, anchor = 'se')

root.title("Receiver collector")

#input folder label, and button

#output folder label, and button

tk.Label(root, text="Input Excel File").grid(row = 1, sticky = 'W')
but_xl = tk.Button(root, text="Input Excel File", fg='blue', command=get_xl_list)
but_xl.grid(row = 1, column = 1, sticky = 'W')

tk.Label(root, text="Output Folder").grid(row = 2, sticky = 'W')
but_path = tk.Button(root, text="Output Folder", fg='blue', command=folder_out_func)
but_path.grid(row = 2, column = 1, sticky = 'W')

but_run = tk.Button(root, text="File", fg='blue', command=run_collection)
but_run.grid(row = 3, column = 1, sticky = 'W')

#print(list_PN_to_get)
#print(output_loc)

#but_path.wait_window()

root.grid_columnconfigure(0, pad=20)
root.grid_columnconfigure(1, pad=20)
root.grid_columnconfigure(2, pad=20)

root.grid_rowconfigure(0, pad=20)
root.grid_rowconfigure(1, pad=20)
root.grid_rowconfigure(2, pad=20)

root.mainloop()
